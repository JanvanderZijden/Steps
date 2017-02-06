from __future__ import division
import pygame, time, os, sys
import RPi.GPIO as GPIO
from gpiozero import MCP3008
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime


def datalogger_init():
	if not os.path.isdir("/home/pi/usbdrv/Beweegdata"):
		os.system("mkdir /home/pi/usbdrv/Beweegdata") 
	if os.path.isfile("/home/pi/usbdrv/Beweegdata/Beweegdata_Steps.xlsx"):
		datalogger = openpyxl.load_workbook("/home/pi/usbdrv/Beweegdata/Beweegdata_Steps.xlsx") #opent het bestaande document op de usb-stick
	else:
		datalogger = openpyxl.load_workbook("/home/pi/Documents/Beweegdata_Steps_Empty.xlsx") #opent een niet ingevuld exceldocument
	return datalogger

def datalogger_sheet():
	dataloggerSheet = datalogger.active
	return dataloggerSheet

def datalogger_save(oefening = "Extensie knie", tijdOefening = "1 min", aantalKeer = 0, afgerond = "ja", moeilijkheid = 1, client = "client"):
	rowCounter = 2
	while dataloggerSheet.cell(row=rowCounter, column=1).value != None:
		rowCounter += 1
		
	dataloggerSheet.cell(row=rowCounter, column=1).value = datetime.now().strftime('%d-%m-%Y')
	dataloggerSheet.cell(row=rowCounter, column=2).value = datetime.now().strftime('%H:%M')
	dataloggerSheet.cell(row=rowCounter, column=3).value = oefening
	dataloggerSheet.cell(row=rowCounter, column=4).value = tijdOefening
	dataloggerSheet.cell(row=rowCounter, column=5).value = aantalKeer
	dataloggerSheet.cell(row=rowCounter, column=6).value = afgerond
	if afgerond == "ja":
		dataloggerSheet.cell(row=rowCounter, column=6).fill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type = 'solid')
	elif afgerond == "nee":
		dataloggerSheet.cell(row=rowCounter, column=6).fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = "solid")
	dataloggerSheet.cell(row=rowCounter, column=7).value = moeilijkheid
	dataloggerSheet.cell(row=rowCounter, column=8).value = client
	datalogger.save("/home/pi/usbdrv/Beweegdata/Beweegdata_Steps.xlsx")
	datalogger.save("/home/pi/Documents/Beweegdata/Beweegdata_Steps.xlsx") #Slaat op als root in niet-startx modus

def convert_timedelta(duration):
	seconds = duration.seconds
	minutes = (seconds % 3600) // 60
	seconds = (seconds % 60)
	return "{} min en {} sec".format(minutes, seconds)

def oefenschema_init():
        if os.path.isfile("/home/pi/usbdrv/Beweegdata/Oefenschema Steps.xlsx"):
                schema = openpyxl.load_workbook("/home/pi/usbdrv/Beweegdata/Oefenschema Steps.xlsx")
                oefenschema = schema.get_sheet_by_name("Oefenschema")
                useOefenschema = True
        else:
                oefenschema = None
                useOefenschema = False
        return oefenschema, useOefenschema

def oefenschema_read(oefenschema):
        gekozenOefeningen = []
        for columnCounter in range(2, 7):
                if oefenschema.cell(row=2, column=columnCounter).value != None: #Als er dus iets in de cell staat.
                        gekozenOefeningen.append(columnCounter)
        return gekozenOefeningen
                        

datalogger = datalogger_init()  #Excel document openen of aanmaken
dataloggerSheet = datalogger_sheet()

oefenschema, useOefenschema = oefenschema_init()
if useOefenschema:
        gekozenOefeningen = oefenschema_read(oefenschema)
        print gekozenOefeningen


GPIO.setmode(GPIO.BCM)
GPIO.setup(4, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)

sensorLinks = MCP3008(1)
sensorRechts = MCP3008(3)
sensorTrigger = 0.02
grote_kaartjes = [[],[],[],[],[],[],[]]
kleine_kaartjes = [None,None,None,None,None,None,None]
try:
	x = (sys.argv[1])
except IndexError:
	x = 0        

pygame.init()
windowInfo = pygame.display.Info()
screen_w = windowInfo.current_w
screen_h = windowInfo.current_h
screen_size = (screen_w,screen_h)
screen = pygame.display.set_mode((screen_w, screen_h))
pygame.mouse.set_visible(False)
#Audio Init
welkom_audio = pygame.mixer.Sound("/home/pi/Documents/Audio/welkom_bij_steps.wav")
goed_gedaan1 = pygame.mixer.Sound("/home/pi/Documents/Audio/goed_gedaan.wav")
goed_gedaan2 = pygame.mixer.Sound("/home/pi/Documents/Audio/10_keer_gedaan_goed_werk.wav")
goed_gedaan3 = pygame.mixer.Sound("/home/pi/Documents/Audio/goed_bezig.wav")
goed_gedaan_special = pygame.mixer.Sound("/home/pi/Documents/Audio/goed_gedaan_barbershop.wav")

achtergrond = pygame.image.load("/home/pi/Documents/Photos/Steps_achtergrond_def.png")
oefening_klaar = pygame.image.load("/home/pi/Documents/Instructie/Oefening_Klaar.png")
achtergrond = pygame.transform.scale(achtergrond, (screen_w, screen_h),)
oefening_klaar = pygame.transform.scale(oefening_klaar, (screen_w, screen_h),)

pygame.mixer.Sound.play(welkom_audio)

def kaartjes_scalen():
	aantalOefeningen = 8
	for file in os.listdir("/home/pi/Documents/Kaartjes"):
		if file.endswith(".png")|file.endswith(".PNG"):
			raw_kaartje = pygame.image.load("/home/pi/Documents/Kaartjes/"+file)
			for oefeningNummer in range(0, aantalOefeningen):
				if file.startswith("%d" % oefeningNummer):
					groot_plaatje = pygame.transform.smoothscale(raw_kaartje,(int (screen_w/2.33),int(pygame.Surface.get_height(raw_kaartje)/pygame.Surface.get_width(raw_kaartje)*(screen_w/2.33))))
					grote_kaartjes[oefeningNummer-1].append(groot_plaatje)
					if (len(grote_kaartjes[oefeningNummer-1]) == 1):
						klein_plaatje = pygame.transform.smoothscale(raw_kaartje,(int (screen_w/4.64),int(pygame.Surface.get_height(raw_kaartje)/pygame.Surface.get_width(raw_kaartje)*(screen_w/4.64))))
						kleine_kaartjes[oefeningNummer-1]=klein_plaatje

##!let op vanaf nu staat oefeningen 1 dus op positie grote_kaartjes[0]. Hierdoor kan oefening 0 dus niet worden toegevoegd. Oefeningen moeten altijd een nummer groter dan 0 hebben.

kaartjes_scalen()
screen.blit(achtergrond, (0,0))

huidige_kaartje_x = ((screen_w/2)- pygame.Surface.get_width(grote_kaartjes[0][0])/2)
huidige_kaartje_y = ((screen_h/2)- pygame.Surface.get_height(grote_kaartjes[0][0])/1.75)
vorige_kaartje_x = ((huidige_kaartje_x/2)-pygame.Surface.get_width(kleine_kaartjes[1])/2)
klein_kaartje_y = ((screen_h/2)- pygame.Surface.get_height(kleine_kaartjes[0])/1.75)
volgende_kaartje_x = (((screen_w-(pygame.Surface.get_width(grote_kaartjes[0][0])+huidige_kaartje_x))/2+(pygame.Surface.get_width(grote_kaartjes[0][0])+huidige_kaartje_x))-pygame.Surface.get_width(kleine_kaartjes[0])/2)

#dit moet 0 zijn omdat oefeningen bij 0 begint
animatie = 0 #deze variabele reguleert de animatie
animatie_counter = 0
sensor_counter = 0
links = 0
rechts = 0
lock_rechts = 0
lock_links = 0
keuze_lock = 0
klaar_counter = 0
while True:
	klaar_counter = 0
	events = pygame.event.get()
	for event in events:
		if event.type == pygame.KEYDOWN:
			pygame.quit()
			sys.exit(0)
	if sensorLinks.value > sensorTrigger and sensorRechts.value < sensorTrigger:
		lock_links = 1
		keuze_lock = 0
	elif sensorRechts.value > sensorTrigger and sensorLinks.value < sensorTrigger:
		lock_rechts = 1
		keuze_lock = 0
	elif sensorRechts.value < sensorTrigger and sensorLinks.value < sensorTrigger:
		keuze_lock = 0

	if (sensorLinks.value < sensorTrigger and lock_links == 1):
		lock_links = 0
		if x > 0:
			x = x - 1
			animatie = 0
	if (sensorRechts.value < sensorTrigger and lock_rechts == 1):
		lock_rechts = 0
		if x < len(kleine_kaartjes)-1:
			x = x + 1
			animatie = 0
	

	if (x-1 >= 0 ):
		screen.blit(kleine_kaartjes[x-1], (vorige_kaartje_x,klein_kaartje_y))
	else:
		screen.blit(achtergrond,(vorige_kaartje_x,klein_kaartje_y),(vorige_kaartje_x,klein_kaartje_y,pygame.Surface.get_width(kleine_kaartjes[0]),pygame.Surface.get_height(kleine_kaartjes[0])))
	if (x+1 < len(kleine_kaartjes)):
		screen.blit(kleine_kaartjes[x+1], (volgende_kaartje_x,klein_kaartje_y))
	else:
		screen.blit(achtergrond, (volgende_kaartje_x,klein_kaartje_y),(volgende_kaartje_x,klein_kaartje_y,pygame.Surface.get_width(kleine_kaartjes[0]),pygame.Surface.get_height(kleine_kaartjes[0])))

	screen.blit(grote_kaartjes[x][animatie], (huidige_kaartje_x,huidige_kaartje_y))
	#animatie
	animatie_counter += 1
	if(animatie_counter > 12):
		animatie_counter = 0
		if (animatie == len(grote_kaartjes[x])-1):
			animatie = 0
		else:
			animatie += 1
	
	#keuze
	if((sensorLinks.value > sensorTrigger and sensorRechts.value > sensorTrigger and keuze_lock == 0) or useOefenschema):
		keuze_lock = 1
		#dirty coding
		if useOefenschema:
			x = gekozenOefeningen.pop()
			if (len(gekozenOefeningen) < 1): #list is leeg
				useOefenschema = False
		#dirty coding
		if(x == 0):
			pygame.quit()
			os.system("python /home/pi/python_code/Steps2/steps_v2_intro_ExtensieKnie.py")
			oefeningBegin = datetime.now()
			os.system("python /home/pi/python_code/Steps2/steps_v2_ExtensieKnie.py")
			tijdOefening = (datetime.now() - oefeningBegin)
			datalogger_save(oefening = "Extensie knie", tijdOefening = convert_timedelta(tijdOefening), aantalKeer = 10)
			pygame.init()
			screen = pygame.display.set_mode((screen_w, screen_h))
			screen.blit(oefening_klaar, (0,0))
			pygame.display.flip()
			pygame.mixer.Sound.play(goed_gedaan2)
			while(klaar_counter < 40):
				time.sleep(0.1)
				if(sensorLinks.value < 0.1 and sensorRechts.value <  0.1):
					klaar_counter += 1
				else:
					klaar_counter = 0
			screen.blit(achtergrond, (0,0))
			lock_links = 0
			lock_rechts = 0
			time.sleep(2)
				
		elif(x == 1):
			pygame.quit()
			os.system("python /home/pi/python_code/Steps2/steps_v2_intro_StaanZitten.py")
			oefeningBegin = datetime.now()
			os.system("python /home/pi/python_code/Steps2/steps_v2_StaanZitten.py")
			tijdOefening = (datetime.now() - oefeningBegin)
			datalogger_save(oefening = "Staan zitten", tijdOefening = convert_timedelta(tijdOefening), aantalKeer = 10)
			pygame.init()
			screen = pygame.display.set_mode((screen_w, screen_h))
			screen.blit(oefening_klaar, (0,0))
			pygame.display.flip()
			screen.blit(achtergrond, (0,0))
			pygame.mixer.Sound.play(goed_gedaan1)
			while(klaar_counter < 40):
				time.sleep(0.1)
				if(sensorLinks.value < 0.1 and sensorRechts.value <  0.1):
					klaar_counter += 1
				else:
					klaar_counter = 0
			lock_links = 0
			lock_rechts = 0
			time.sleep(2)
				
		elif(x == 2):
			pygame.quit()
			os.system("python /home/pi/python_code/Steps2/steps_v2_intro_AchterenLopen.py")
			oefeningBegin = datetime.now()
			os.system("python /home/pi/python_code/Steps2/steps_v2_AchterenLopen.py")
			tijdOefening = (datetime.now() - oefeningBegin)
			datalogger_save(oefening = "Naar achter lopen", tijdOefening = convert_timedelta(tijdOefening), aantalKeer = 10)
			pygame.init()
			screen = pygame.display.set_mode((screen_w, screen_h))
			screen.blit(oefening_klaar, (0,0))
			pygame.display.flip()
			screen.blit(achtergrond, (0,0))
			pygame.mixer.Sound.play(goed_gedaan3)
			while(klaar_counter < 40):
				time.sleep(0.1)
				if(sensorLinks.value < 0.1 and sensorRechts.value <  0.1):
					klaar_counter += 1
				else:
					klaar_counter = 0
			lock_rechts = 0
			lock_links = 0
			time.sleep(2)
				
		elif(x == 3):
			pygame.quit()
			os.system("python /home/pi/python_code/Steps2/steps_v2_intro_BeenHeffen.py")
			oefeningBegin = datetime.now()
			os.system("python /home/pi/python_code/Steps2/steps_v2_BeenHeffen.py")
			tijdOefening = (datetime.now() - oefeningBegin)
			datalogger_save(oefening = "Been heffen", tijdOefening = convert_timedelta(tijdOefening), aantalKeer = 10)
			pygame.init()
			screen = pygame.display.set_mode((screen_w, screen_h))
			screen.blit(oefening_klaar, (0,0))
			pygame.display.flip()
			screen.blit(achtergrond, (0,0))
			pygame.mixer.Sound.play(goed_gedaan2)
			while(klaar_counter < 40):
				time.sleep(0.1)
				if(sensorLinks.value < 0.1 and sensorRechts.value <  0.1):
					klaar_counter += 1
				else:
					klaar_counter = 0
			lock_links = 0
			lock_rechts = 0
			time.sleep(2)
				
		elif(x == 4):
			pygame.quit()
			os.system("python /home/pi/python_code/Steps2/steps_v2_intro_HakNaarBil.py")
			oefeningBegin = datetime.now()
			os.system("python /home/pi/python_code/Steps2/steps_v2_HakNaarBil.py")
			tijdOefening = (datetime.now() - oefeningBegin)
			datalogger_save(oefening = "Hak naar bil", tijdOefening = convert_timedelta(tijdOefening), aantalKeer = 10)
			pygame.init()
			screen = pygame.display.set_mode((screen_w, screen_h))
			screen.blit(oefening_klaar, (0,0))
			pygame.display.flip()
			screen.blit(achtergrond, (0,0))
			pygame.mixer.Sound.play(goed_gedaan3)
			while(klaar_counter < 40):
				time.sleep(0.1)
				if(sensorLinks.value < 0.1 and sensorRechts.value <  0.1):
					klaar_counter += 1
				else:
					klaar_counter = 0
			lock_links = 0
			lock_rechts = 0
			time.sleep(2)
				
		elif(x == 5):
			pygame.quit()
			os.system("python /home/pi/python_code/Steps2/steps_v2_intro_StaanOpEenBeen.py")
			oefeningBegin = datetime.now()
			os.system("python /home/pi/python_code/Steps2/steps_v2_OpEenBeenStaanMs.py")
			tijdOefening = (datetime.now() - oefeningBegin)
			datalogger_save(oefening = "Staan op een been met steun", tijdOefening = convert_timedelta(tijdOefening), aantalKeer = 10)
			pygame.init()
			screen = pygame.display.set_mode((screen_w, screen_h))
			screen.blit(oefening_klaar, (0,0))
			pygame.display.flip()
			screen.blit(achtergrond, (0,0))
			pygame.mixer.Sound.play(goed_gedaan2)
			while(klaar_counter < 40):
				time.sleep(0.1)
				if(sensorLinks.value < 0.1 and sensorRechts.value <  0.1):
					klaar_counter += 1
				else:
					klaar_counter = 0
			lock_links = 0
			lock_rechts = 0
			time.sleep(2)
				
		elif(x == 6):
			pygame.quit()
			os.system("python /home/pi/python_code/Steps2/steps_v2_intro_StaanOpEenBeen.py")
			oefeningBegin = datetime.now()
			os.system("python /home/pi/python_code/Steps2/steps_v2_OpEenBeenStaanZs.py")
			tijdOefening = (datetime.now() - oefeningBegin)
			datalogger_save(oefening = "Staan op een been zonder steun", tijdOefening = convert_timedelta(tijdOefening), aantalKeer = 10)
			pygame.init()
			screen = pygame.display.set_mode((screen_w, screen_h))
			screen.blit(oefening_klaar, (0,0))
			pygame.display.flip()
			screen.blit(achtergrond, (0,0))
			pygame.mixer.Sound.play(goed_gedaan_special)
			while(klaar_counter < 40):
				time.sleep(0.1)
				if(sensorLinks.value < 0.1 and sensorRechts.value <  0.1):
					klaar_counter += 1
				else:
					klaar_counter = 0
			lock_links = 0
			lock_rechts = 0
			time.sleep(2)
	pygame.display.flip()


	
