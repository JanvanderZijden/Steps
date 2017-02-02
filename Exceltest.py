import time
import pygame
import sys
import openpyxl

schema = openpyxl.load_workbook("/home/pi/usbdrv/Oefendata/Oefenschema Steps.xlsx")
oefenschema = schema.get_sheet_by_name("Oefenschema")

datalogger = openpyxl.Workbook()
dataloggerSheet = datalogger.active
dataloggerSheet.title = "Data"
dataloggerSheet.cell(row=2, column=2).value = "10 uur"
datalogger.save("/home/pi/usbdrv/Oefendata/Datalogger Excel Steps.xlsx")

pygame.init()
pygame.font.init()
screen = pygame.display.set_mode()
windowInfo = pygame.display.Info()

myfont = pygame.font.SysFont("Comic Sans MS", 30)

while True:
	pygame.draw.rect(screen,(0,0,0),(0,0,windowInfo.current_w,windowInfo.current_h))

	text01 = myfont.render(str(oefenschema.cell(row=2,column=2).value + 1), 1, (255,255,0))
	
	screen.blit(text01,(710,windowInfo.current_h-100))
	
	pygame.display.flip()
	events = pygame.event.get()
	for event in events:
		if event.type == pygame.KEYDOWN:
			pygame.quit()
			sys.exit(1)
	time.sleep(0.1)
